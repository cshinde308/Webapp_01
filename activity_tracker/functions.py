import streamlit as st
from datetime import datetime
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
import streamlit.components.v1 as components
from datetime import date, timedelta


# ================= HELPERS =================

def render_preview_with_multiline_notes1(df, height_px=400):
    """
    Render dataframe as scrollable HTML table using components.html()
    (guaranteed rendering).
    """
    df = df.copy()

    # Convert newlines to <br>
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].fillna("").astype(str).str.replace("\n", "<br>")

    html_table = df.to_html(index=False, escape=False)

    html = f"""
    <style>
        .scroll-table {{
            max-height: {height_px}px;
            overflow-y: auto;
            border: 1px solid #ccc;
            font-family: Arial, sans-serif;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 6px;
            border: 1px solid #ddd;
            text-align: left;
            vertical-align: top;
            font-size: 13px;
        }}
        td {{
            white-space: pre-wrap;
            word-wrap: break-word;
        }}
        thread th {{
            position: sticky;
            top: 0;
            background-color: #f2f2f2;
            z-index: 10;
        }}
    </style>

    <div class="scroll-table">
        {html_table}
    </div>
    """

    components.html(html, height=height_px + 60, scrolling=True)

def render_preview_with_multiline_notes(df, height_px=400):
    """
    Render dataframe as scrollable HTML table using components.html()
    (Streamlit Cloud safe).
    """
    df = df.copy()

    # ✅ FIX: handle Windows + Linux newlines
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = (
                df[col]
                .fillna("")
                .astype(str)
                .str.replace(r"\r?\n", "<br>", regex=True)
            )

    html_table = df.to_html(index=False, escape=False)

    html = f"""
    <style>
        .scroll-table {{
            max-height: {height_px}px;
            overflow-y: auto;
            border: 1px solid #ccc;
            font-family: Arial, sans-serif;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 6px;
            border: 1px solid #ddd;
            text-align: left;
            vertical-align: top;
            font-size: 13px;
        }}
        td {{
            white-space: normal;   /* IMPORTANT */
            word-wrap: break-word;
        }}
        thead th {{
            position: sticky;
            top: 0;
            background-color: #f2f2f2;
            z-index: 10;
        }}
    </style>

    <div class="scroll-table">
        {html_table}
    </div>
    """

    components.html(html, height=height_px + 60, scrolling=True)

def parse_excel_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        try:
            return datetime.strptime(val, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

def create_weekly_summary_excel(records, summary_type):
    # ✅ SORT ALL RECORDS BY DATE THEN TIME
    records = sorted(
        records,
        key=lambda x: (x["date"], x["time"])
    )

    wb = Workbook()
    bold_font = Font(bold=True)

    # =============================
    # DATE‑WISE (Single sheet)
    # =============================
    if summary_type == "Date-wise":
        ws = wb.active
        ws.title = "Date-wise Summary"
        ws.append(["Date", "Time", "Project", "Engineer", "Hours", "Note"])

        total_hours = 0.0

        for r in records:
            ws.append([
                r["date"].strftime("%Y-%m-%d"),
                r["time"],
                r["project"],
                r["engineer"],
                r["hours"],
                r["note"]
            ])
            ws[f"F{ws.max_row}"].alignment = Alignment(wrap_text=True)
            total_hours += r["hours"]

        # ✅ TOTAL ROW (BOLD)
        ws.append(["", "", "Total:", "", total_hours, ""])
        ws[f"C{ws.max_row}"].font = bold_font
        ws[f"E{ws.max_row}"].font = bold_font

    # =============================
    # ENGINEER‑WISE
    # =============================
    elif summary_type == "Engineer-wise":
        grouped = {}

        for r in records:
            grouped.setdefault(r["engineer"], []).append(r)

        for engineer, items in grouped.items():
            ws = wb.create_sheet(title=engineer[:31])
            ws.append(["Date", "Time", "Project", "Engineer", "Hours", "Note"])

            total_hours = 0.0

            for r in items:
                ws.append([
                    r["date"].strftime("%Y-%m-%d"),
                    r["time"],
                    r["project"],
                    r["engineer"],
                    r["hours"],
                    r["note"]
                ])
                ws[f"F{ws.max_row}"].alignment = Alignment(wrap_text=True)
                total_hours += r["hours"]

            # ✅ TOTAL ROW (BOLD)
            ws.append(["", "", "Total:", "", total_hours, ""])
            ws[f"C{ws.max_row}"].font = bold_font
            ws[f"E{ws.max_row}"].font = bold_font

    # =============================
    # PROJECT‑WISE
    # =============================
    elif summary_type == "Project-wise":
        grouped = {}

        for r in records:
            grouped.setdefault(r["project"], []).append(r)

        for project, items in grouped.items():
            ws = wb.create_sheet(title=project[:31])
            ws.append(["Date", "Time", "Project", "Engineer", "Hours", "Note"])

            total_hours = 0.0

            for r in items:
                ws.append([
                    r["date"].strftime("%Y-%m-%d"),
                    r["time"],
                    r["project"],
                    r["engineer"],
                    r["hours"],
                    r["note"]
                ])
                ws[f"F{ws.max_row}"].alignment = Alignment(wrap_text=True)
                total_hours += r["hours"]

            # ✅ TOTAL ROW (BOLD)
            ws.append(["", "", "Total:", "", total_hours, ""])
            ws[f"C{ws.max_row}"].font = bold_font
            ws[f"E{ws.max_row}"].font = bold_font

    # =============================
    # REMOVE DEFAULT EMPTY SHEET
    # =============================
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    return wb


def append_weekly_summary(records,excel_file):
    wb = load_workbook(excel_file)
    name = f"Weekly_Summary_{datetime.now():%Y%m%d_%H%M}"
    ws = wb.create_sheet(name)
    ws.append(["Date", "Time", "Project", "Engineer", "Hours", "Note"])

    for r in records:
        ws.append([
            r["date"].strftime("%Y-%m-%d"),
            r["time"],
            r["project"],
            r["engineer"],
            r["hours"],
            r["note"]
        ])
        ws[f"F{ws.max_row}"].alignment = Alignment(wrap_text=True)

    wb.save(excel_file)
    wb.close()

def get_activity_block_html(target_date,excel_file):
    wb = load_workbook(excel_file, read_only=True)
    ws = wb["Log"]

    lines = []
    total_hours = 0.0

    for r in ws.iter_rows(min_row=2, values_only=True):

        # ✅ ROLE‑BASED VISIBILITY FILTER
        if st.session_state.role != "Administrator":
            if r[3] != st.session_state.username:
                continue

        d = parse_excel_date(r[0])
        if d == target_date:
            time = r[1] or "--:--"
            project = r[2] or "-"
            engineer = r[3] or "-"
            hours = float(r[4]) if r[4] else 0.0

            # ✅ IMPORTANT FIX HERE
            note = (r[5] or "").replace("\n", "<br>")

            total_hours += hours

            lines.append(
                f"<b>{time}</b> | {project} | {engineer} | <b>{hours} h</b><br>"
                f"<span style='color:#666'>{note}</span>"
            )

    wb.close()

    if not lines:
        content = "<i>No activities logged.</i>"
    else:
        content = "<hr style='margin:6px 0'>".join(lines)

    html = f"""
    <div style="
        font-size:12px;
        line-height:1.4;
        max-height:220px;
        overflow-y:auto;
        border:1px solid #ddd;
        padding:8px;
        background-color:#fafafa;
        border-radius:4px;
    ">
        <div style="font-weight:600; margin-bottom:6px;">
            Total Hours: {total_hours:.2f}
        </div>
        {content}
    </div>
    """
    return html

def render_time_entry_code_view(records):
    """
    Renders a compact text-style view similar to external booking tools.
    """
    grouped = {}

    for r in records:
        grouped.setdefault(r["project"], 0.0)
        grouped[r["project"]] += r["hours"]

    total_hours = sum(grouped.values())

    lines = []
    for proj, hrs in grouped.items():
        lines.append(f"{proj} : {hrs:.2f}")

    text = "\n".join(lines)

    st.text_area(
        "Time Entry Code View",
        value=text + f"\n\nTotal Hours = {total_hours:.2f}",
        height=300,
        disabled=True
    )

def render_time_entry_code_view_html(records, proj_map):
    """
    Larger, visual booking-tool style view
    showing Project + IFS code + total hours
    """
    project_hours = {}

    for r in records:
        project_hours.setdefault(r["project"], 0.0)
        project_hours[r["project"]] += r["hours"]

    total_hours = sum(project_hours.values())

    rows_html = ""
    for project, hrs in sorted(project_hours.items()):
        ifs_code = proj_map.get(project, "")
        ifs_text = f" ({ifs_code})" if ifs_code else ""

        rows_html += f"""
        <div style="
            display:flex;
            justify-content:space-between;
            padding:6px 0;
            font-size:15px;
        ">
            <span style="font-weight:600;">
                {project}{ifs_text}
            </span>
            <span style="font-weight:600;">
                {hrs:.2f} h
            </span>
        </div>
        """

    html = f"""
    <div style="
        font-size:15px;
        line-height:1.6;
        max-height:360px;
        overflow-y:auto;
        border:1px solid #d0d0d0;
        padding:14px;
        background-color:#f9fbfc;
        border-radius:8px;
    ">
        <div style="
            font-size:16px;
            font-weight:700;
            margin-bottom:10px;
        ">
            ⏱ Time Entry Code View
        </div>

        {rows_html if rows_html else "<i>No data available</i>"}

        <hr style="margin:12px 0;">

        <div style="
            font-size:16px;
            font-weight:700;
            text-align:right;
        ">
            Total Hours : {total_hours:.2f}
        </div>
    </div>
    """
    return html

def prev_day():
    st.session_state.activity_context_date -= timedelta(days=1)

def next_day():
    st.session_state.activity_context_date += timedelta(days=1)

def init_session():
    defaults = {
        "logged_in": False,
        "username": None,
        "role": None,
        "last_active": None
    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)

def check_idle_timeout(settings):
    if not st.session_state.logged_in:
        return

    idle_limit = settings.get("popup_min", 10) * 60  # minutes → seconds
    now = datetime.now()

    if st.session_state.last_active:
        idle_time = (now - st.session_state.last_active).total_seconds()
        if idle_time > idle_limit:
            st.session_state.logged_in = False
            st.session_state.username = None
            st.session_state.role = None
            st.session_state.last_active = None
            st.warning("⏱ Session expired. Please login again.")
            st.rerun()

def sidebar_item(icon, text):
    return f'''
    <div class="sidebar-item">
        {icon}
        <span class="sidebar-text">{text}</span>
    </div>
    '''

# ================= WIZARDS =================
def manage_projects_wizard(project_file, projects):
    st.subheader("🛠 Manage Projects (IFS)")

    # ✅ Display dataframe with sentence‑case headers
    df = pd.DataFrame(projects).rename(columns={
        "name": "Name",
        "ifs": "IFS Code"
    })

    edited = st.data_editor(
        df,
        num_rows="dynamic",
        width="stretch"
    )

    # ✅ Save button
    save_projects_clicked = st.button("💾 Save Projects")

    # ✅ Placeholder JUST BELOW button
    projects_msg_placeholder = st.empty()

    if save_projects_clicked:
        # ✅ Convert headers BACK before saving
        edited = edited.rename(columns={
            "Name": "name",
            "IFS Code": "ifs"
        })

        clean = (
            edited
            .dropna(subset=["name"])
            .drop_duplicates(subset=["name"])
            .to_dict("records")
        )

        json.dump(clean, open(project_file, "w"), indent=2)

        # ✅ Set flash flag
        st.session_state.projects_saved = True
        st.rerun()

    # ✅ ONE‑TIME success message (FLASH) – BELOW button
    if st.session_state.pop("projects_saved", False):
        projects_msg_placeholder.success("✅ Projects saved successfully")

def manage_engineers_wizard(engineers_file, engineers):
    st.subheader("👷 Manage Engineers")

    df = pd.DataFrame(engineers, columns=["Engineer"])

    edited = st.data_editor(
        df,
        num_rows="dynamic",
        width="stretch"
    )

    # ✅ Save button
    save_engineers_clicked = st.button("💾 Save Engineers")

    # ✅ Placeholder JUST BELOW button
    engineers_msg_placeholder = st.empty()

    if save_engineers_clicked:
        clean = (
            edited["Engineer"]
            .dropna()
            .drop_duplicates()
            .tolist()
        )

        json.dump(clean, open(engineers_file, "w"), indent=2)

        # ✅ Set flash flag
        st.session_state.engineers_saved = True
        st.rerun()

    # ✅ ONE‑TIME success message (FLASH) – BELOW button
    if st.session_state.pop("engineers_saved", False):
        engineers_msg_placeholder.success("✅ Engineers saved successfully")


