import streamlit as st
from datetime import datetime, timedelta, date
from pathlib import Path
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from io import BytesIO
import streamlit.components.v1 as components
from datetime import date, timedelta

# ✅ Sidebar expand control
if "_force_sidebar_open" not in st.session_state:
    st.session_state["_force_sidebar_open"] = False
# ================= PAGE CONFIG =================
st.set_page_config(
    page_title="Activity Tracker",
    page_icon="🕒",
    layout="wide",
    initial_sidebar_state="expanded"
    if st.session_state["_force_sidebar_open"]
    else "collapsed"
)
##==================SIDEBAR CSS====================================##
st.markdown("""
<style>
div[data-testid="stSidebarContent"] {
    display: flex;
    flex-direction: column;
    height: 100vh;
}

.sidebar-bottom {
    margin-top: auto;
    padding: 12px;
    border-top: 1px solid #e5e7eb;
}

.user-card {
    background: #ffffff;
    border-radius: 10px;
    padding: 12px;
    box-shadow: 0 2px 6px rgba(0,0,0,0.1);
    font-size: 13px;
}
</style>
""", unsafe_allow_html=True)




# ================= PATHS =================
DATA_DIR1 = Path("data")

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

DATA_DIR.mkdir(exist_ok=True)

EXCEL_FILE = DATA_DIR1 / "activity_log.xlsx"
PROJECTS_FILE = DATA_DIR1 / "projects.json"
ENGINEERS_FILE = DATA_DIR1 / "Engineers.json"
SETTINGS_FILE = DATA_DIR1 / "settings.json"

# ================= DEFAULT DATA =================
DEFAULT_PROJECTS = [
    {"name": "Project-1", "ifs": ""},
    {"name": "Project-2", "ifs": ""},
    {"name": "Project-3", "ifs": ""},
]

DEFAULT_ENGINEERS = [
    "Engineer-1", "Engineer-2", "Engineer-3",
    "Engineer-4", "Engineer-5",
]

# ================= FILE INITIALIZATION =================
def ensure_files():
    if not EXCEL_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(["Date", "Time", "Project", "Engineer", "Hours", "Note"])
        wb.save(EXCEL_FILE)
        wb.close()

    if not PROJECTS_FILE.exists():
        json.dump(DEFAULT_PROJECTS, open(PROJECTS_FILE, "w"), indent=2)

    if not ENGINEERS_FILE.exists():
        json.dump(DEFAULT_ENGINEERS, open(ENGINEERS_FILE, "w"), indent=2)

    if not SETTINGS_FILE.exists():
        json.dump({"popup_min": 5}, open(SETTINGS_FILE, "w"), indent=2)

ensure_files()


USERS_FILE = DATA_DIR / "users.json"

if not USERS_FILE.exists():
    st.error("❌ users.json file not found in data folder")
    st.stop()

with open(USERS_FILE, "r") as f:
    users_db = json.load(f)


# ✅ TEMP DEBUG (REMOVE AFTER TEST)
#st.write("DEBUG: Loaded users =", users_db)
#st.write("✅ users.json path:", USERS_FILE.resolve())
#st.write("✅ users.json content:", users_db)





# ================= LOAD DATA =================
projects = json.load(open(PROJECTS_FILE))
engineers = json.load(open(ENGINEERS_FILE))
settings = json.load(open(SETTINGS_FILE))

proj_names = [p["name"] for p in projects]
proj_map = {p["name"]: p.get("ifs", "") for p in projects}

# ================= HELPERS =================

def render_preview_with_multiline_notes(df, height_px=400):
    """
    Render dataframe as scrollable HTML table using components.html()
    (guaranteed rendering).
    """
    df = df.copy()

    # Convert newlines to <br>
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].fillna("").astype(str).str.replace("\n", "<br>")

    html_table = df.to_html(index=False, escape=False)

    html = f"""
    <style>
        .scroll-table {{
            max-height: {height_px}px;
            overflow-y: auto;
            border: 1px solid #ccc;
            font-family: Arial, sans-serif;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 6px;
            border: 1px solid #ddd;
            text-align: left;
            vertical-align: top;
            font-size: 13px;
        }}
        td {{
            white-space: pre-wrap;
            word-wrap: break-word;
        }}
        thead th {{
            position: sticky;
            top: 0;
            background-color: #f2f2f2;
            z-index: 10;
        }}
    </style>

    <div class="scroll-table">
        {html_table}
    </div>
    """

    components.html(html, height=height_px + 60, scrolling=True)


def parse_excel_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        try:
            return datetime.strptime(val, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

def create_weekly_summary_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Summary"
    ws.append(["Date", "Time", "Project", "Engineer", "Hours", "Note"])

    for r in records:
        ws.append([
            r["date"].strftime("%Y-%m-%d"),
            r["time"],
            r["project"],
            r["engineer"],
            r["hours"],
            r["note"]
        ])
        ws[f"F{ws.max_row}"].alignment = Alignment(wrap_text=True)

    return wb

def append_weekly_summary(records):
    wb = load_workbook(EXCEL_FILE)
    name = f"Weekly_Summary_{datetime.now():%Y%m%d_%H%M}"
    ws = wb.create_sheet(name)
    ws.append(["Date", "Time", "Project", "Engineer", "Hours", "Note"])

    for r in records:
        ws.append([
            r["date"].strftime("%Y-%m-%d"),
            r["time"],
            r["project"],
            r["engineer"],
            r["hours"],
            r["note"]
        ])
        ws[f"F{ws.max_row}"].alignment = Alignment(wrap_text=True)

    wb.save(EXCEL_FILE)
    wb.close()

def get_activity_block_html(target_date):
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb["Log"]

    lines = []
    total_hours = 0.0

    for r in ws.iter_rows(min_row=2, values_only=True):
        d = parse_excel_date(r[0])
        if d == target_date:
            time = r[1] or "--:--"
            project = r[2] or "-"
            engineer = r[3] or "-"
            hours = float(r[4]) if r[4] else 0.0

            # ✅ IMPORTANT FIX HERE
            note = (r[5] or "").replace("\n", "<br>")

            total_hours += hours

            lines.append(
                f"<b>{time}</b> | {project} | {engineer} | <b>{hours} h</b><br>"
                f"<span style='color:#666'>{note}</span>"
            )

    wb.close()

    if not lines:
        content = "<i>No activities logged.</i>"
    else:
        content = "<hr style='margin:6px 0'>".join(lines)

    html = f"""
    <div style="
        font-size:12px;
        line-height:1.4;
        max-height:220px;
        overflow-y:auto;
        border:1px solid #ddd;
        padding:8px;
        background-color:#fafafa;
        border-radius:4px;
    ">
        <div style="font-weight:600; margin-bottom:6px;">
            Total Hours: {total_hours:.2f}
        </div>
        {content}
    </div>
    """
    return html

def render_time_entry_code_view(records):
    """
    Renders a compact text-style view similar to external booking tools.
    """
    grouped = {}

    for r in records:
        grouped.setdefault(r["project"], 0.0)
        grouped[r["project"]] += r["hours"]

    total_hours = sum(grouped.values())

    lines = []
    for proj, hrs in grouped.items():
        lines.append(f"{proj} : {hrs:.2f}")

    text = "\n".join(lines)

    st.text_area(
        "Time Entry Code View",
        value=text + f"\n\nTotal Hours = {total_hours:.2f}",
        height=300,
        disabled=True
    )

def render_time_entry_code_view_html(records, proj_map):
    """
    Larger, visual booking-tool style view
    showing Project + IFS code + total hours
    """
    project_hours = {}

    for r in records:
        project_hours.setdefault(r["project"], 0.0)
        project_hours[r["project"]] += r["hours"]

    total_hours = sum(project_hours.values())

    rows_html = ""
    for project, hrs in sorted(project_hours.items()):
        ifs_code = proj_map.get(project, "")
        ifs_text = f" ({ifs_code})" if ifs_code else ""

        rows_html += f"""
        <div style="
            display:flex;
            justify-content:space-between;
            padding:6px 0;
            font-size:15px;
        ">
            <span style="font-weight:600;">
                {project}{ifs_text}
            </span>
            <span style="font-weight:600;">
                {hrs:.2f} h
            </span>
        </div>
        """

    html = f"""
    <div style="
        font-size:15px;
        line-height:1.6;
        max-height:360px;
        overflow-y:auto;
        border:1px solid #d0d0d0;
        padding:14px;
        background-color:#f9fbfc;
        border-radius:8px;
    ">
        <div style="
            font-size:16px;
            font-weight:700;
            margin-bottom:10px;
        ">
            ⏱ Time Entry Code View
        </div>

        {rows_html if rows_html else "<i>No data available</i>"}

        <hr style="margin:12px 0;">

        <div style="
            font-size:16px;
            font-weight:700;
            text-align:right;
        ">
            Total Hours : {total_hours:.2f}
        </div>
    </div>
    """
    return html

def prev_day():
    st.session_state.activity_context_date -= timedelta(days=1)

def next_day():
    st.session_state.activity_context_date += timedelta(days=1)

def init_session():
    defaults = {
        "logged_in": False,
        "username": None,
        "role": None,
        "last_active": None
    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)

def check_idle_timeout():
    if not st.session_state.logged_in:
        return

    idle_limit = settings.get("popup_min", 10) * 60  # minutes → seconds
    now = datetime.now()

    if st.session_state.last_active:
        idle_time = (now - st.session_state.last_active).total_seconds()
        if idle_time > idle_limit:
            st.session_state.logged_in = False
            st.session_state.username = None
            st.session_state.role = None
            st.session_state.last_active = None
            st.warning("⏱ Session expired. Please login again.")
            st.rerun()

def sidebar_item(icon, text):
    return f'''
    <div class="sidebar-item">
        {icon}
        <span class="sidebar-text">{text}</span>
    </div>
    '''

# ================= WIZARDS =================
def manage_projects_wizard():
    st.subheader("🛠 Manage Projects (IFS)")
    df = pd.DataFrame(projects)

    edited = st.data_editor(
        df,
        num_rows="dynamic",
        width="stretch"
    )

    if st.button("💾 Save Projects"):
        clean = (
            edited.dropna(subset=["name"])
            .drop_duplicates(subset=["name"])
            .to_dict("records")
        )
        json.dump(clean, open(PROJECTS_FILE, "w"), indent=2)
        st.success("Projects saved")
        st.rerun()

def manage_engineers_wizard():
    st.subheader("👷 Manage Engineers")
    df = pd.DataFrame(engineers, columns=["Engineer"])

    edited = st.data_editor(
        df,
        num_rows="dynamic",
        width="stretch"
    )

    if st.button("💾 Save Engineers"):
        clean = (
            edited["Engineer"]
            .dropna()
            .drop_duplicates()
            .tolist()
        )
        json.dump(clean, open(ENGINEERS_FILE, "w"), indent=2)
        st.success("Engineers saved")
        st.rerun()

# ---------------- SESSION INIT ----------------
init_session()

# ---------------- IDLE TIMEOUT CHECK ----------------
check_idle_timeout()

# ✅ STEP‑5 MUST COME HERE
if st.session_state.logged_in:
    st.session_state.last_active = datetime.now()

st.markdown("""
<style>
/* Login box wrapper */
.login-box {
    max-width: 380px;       /* ✅ IMPORTANT: controls form width */
    margin: 80px auto;
}

/* Login form styling ONLY */
.login-box div[data-testid="stForm"] {
    padding: 24px;
    border-radius: 12px;
    background-color: white;
    box-shadow: 0 6px 18px rgba(0,0,0,0.15);
}

/* Center login title */
.login-title {
    text-align: center;
    font-size: 22px;
    font-weight: 600;
    margin-bottom: 16px;
}
</style>
""", unsafe_allow_html=True)



# ---------------- LOGIN SCREEN ----------------
if not st.session_state.logged_in:
    if not st.session_state.logged_in:

        st.markdown('<div class="login-box">', unsafe_allow_html=True)

        st.markdown('<div class="login-title">🔐 Login</div>', unsafe_allow_html=True)

        with st.form("login_form"):
            username_input = st.text_input("Username")
            password_input = st.text_input("Password", type="password")
            login_btn = st.form_submit_button("Login", use_container_width=True)

        if login_btn:
            username = username_input.strip()
            password = password_input.strip()

            # ✅ DEBUG
            st.write("Trying username:", username)

            user = users_db.get(username)

            if user is None:
                st.error("❌ User not found")

            elif user.get("locked"):
                st.error("🔒 Account is locked. Contact administrator.")

            elif user["password"] != password:
                st.error("❌ Invalid password")

            else:
                st.session_state.logged_in = True
                st.session_state.username = username
                st.session_state.role = user["role"]
                st.session_state.last_active = datetime.now()
                st.success("✅ Login successful")
                st.rerun()

        st.stop()

# ================= SIDEBAR =================

# ---- TOP SECTION ----
st.sidebar.markdown("### CS_Toolbox")
st.sidebar.caption("📁 Tools")

page = st.sidebar.radio(
    "Tools",
    ["Activity Tracker", "Page-2"],
    key="sidebar_tools_radio",
    label_visibility="collapsed"
)

# ✅ LOGOUT stays in TOP section
if st.sidebar.button("🚪 Logout", use_container_width=True):
    st.session_state.logged_in = False
    st.session_state.username = None
    st.session_state.role = None
    st.session_state.last_active = None
    st.rerun()

# ---- BOTTOM SECTION (HTML ONLY) ----
st.sidebar.markdown(
    f"""
    <div class="sidebar-bottom">
        <div class="user-card">
            👤 <b>{st.session_state.username}</b><br>
            <span style="color:#6b7280;font-size:12px;">
                {st.session_state.role}
            </span><br>
            <span style="color:#16a34a;font-size:11px;">● Online</span>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)


# ================= MAIN PAGE =================
if page == "Activity Tracker":
    if st.session_state.role == "Administrator":
        tabs = st.tabs(
            ["📊 Project Dashboard", "📝 Log Activity",
             "📅 Edit Activity", "📊 Summary", "⚙ Settings"]
        )
    else:
        tabs = st.tabs(
            ["📊 Project Dashboard", "📝 Log Activity",
             "📅 Edit Activity", "📊 Summary"]
        )

    tab_project_dashboard = tabs[0]
    tab_log = tabs[1]
    tab_editactivity = tabs[2]
    tab_summary = tabs[3]

    if st.session_state.role == "Administrator":
        tab_settings = tabs[4]

    # ---------- PROJECT DASHBOARD ----------
    with tab_project_dashboard:
        st.subheader("📊 Project Dashboard")

        # -----------------------------
        # Load log data
        # -----------------------------
        wb = load_workbook(EXCEL_FILE, read_only=True)
        ws = wb["Log"]

        data = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = parse_excel_date(r[0])
            if not d:
                continue

            data.append({
                "date": d,
                "project": r[2],
                "engineer": r[3],
                "hours": float(r[4]) if r[4] else 0.0
            })

        wb.close()

        if not data:
            st.info("No activity data available.")
        else:
            df = pd.DataFrame(data)

            # -----------------------------
            # FILTERS (Date + Project only)
            # -----------------------------
            col1, col2, col3 = st.columns(3)

            start_date = col1.date_input(
                "Start Date", df["date"].min()
            )

            end_date = col2.date_input(
                "End Date", df["date"].max()
            )

            selected_project = col3.selectbox(
                "Select Project",
                sorted(df["project"].unique())
            )

            # Apply filters
            df = df[
                (df["date"] >= start_date) &
                (df["date"] <= end_date) &
                (df["project"] == selected_project)
                ]

            st.divider()

            if df.empty:
                st.warning("No data found for selected project and date range.")
            else:
                # -----------------------------
                # KPI SUMMARY
                # -----------------------------
                total_hours = df["hours"].sum()
                engineer_count = df["engineer"].nunique()

                c1, c2 = st.columns(2)
                c1.metric("🕒 Total Hours Booked", f"{total_hours:.2f} h")
                c2.metric("👷 Engineers Involved", engineer_count)

                st.divider()

                # -----------------------------
                # ENGINEER CONTRIBUTION
                # -----------------------------

                eng_hours = (
                    df.groupby("engineer", as_index=False)["hours"]
                    .sum()
                    .sort_values("hours", ascending=False)
                )

                # Pie Chart
                st.subheader("🍰 Engineers Contribution (Pie Chart)")

                pie_df = eng_hours.set_index("engineer")

                st.plotly_chart(
                    {
                        "data": [{
                            "labels": pie_df.index,
                            "values": pie_df["hours"],
                            "type": "pie",
                            "hole": 0.4,
                            "textinfo": "label+percent",
                            "textposition": "inside"
                        }],
                        "layout": {
                            "title": "Engineer Contribution by Hours",
                            "showlegend": True,
                            "height": 1000
                        }
                    },
                    width="stretch"
                )

                st.divider()

                # -----------------------------
                # ENGINEER CONTRIBUTION TABLE
                # -----------------------------
                eng_hours["Contribution %"] = (
                        eng_hours["hours"] / total_hours * 100
                ).round(1)

                eng_hours.rename(
                    columns={
                        "engineer": "Engineer",
                        "hours": "Hours"
                    },
                    inplace=True
                )

                st.subheader("📋 Engineer-wise Breakdown")
                st.dataframe(
                    eng_hours,
                    width="stretch",
                    hide_index=True
                )

                st.caption(
                    "This dashboard shows project‑specific effort distribution "
                    "based on selected date range."
                )

    # ---------- LOG ----------
    with tab_log:
        st.subheader("🗓 Activity Context View")

        # ✅ 1. Initialize session state ONCE
        if "activity_context_date" not in st.session_state:
            st.session_state.activity_context_date = date.today()


        # ✅ 2. Arrow callbacks (ONLY place where date changes)
        def prev_day():
            st.session_state.activity_context_date -= timedelta(days=1)


        def next_day():
            st.session_state.activity_context_date += timedelta(days=1)


        # ✅ 3. Date navigation UI
        col_y, col_t, col_tm = st.columns([1, 1, 1])

        with col_t:
            nav_l, nav_m, nav_r = st.columns([1, 4, 1])

            with nav_l:
                st.button("⬅️", on_click=prev_day)

            with nav_m:
                st.date_input(
                    "Select Date",
                    key="activity_context_date"  # ✅ NO value parameter
                )

            with nav_r:
                st.button("➡️", on_click=next_day)

        # ✅ 4. Read date AFTER widget
        base_date = st.session_state.activity_context_date
        yesterday = base_date - timedelta(days=1)
        today = base_date
        tomorrow = base_date + timedelta(days=1)

        # ---- YESTERDAY ----
        with col_y:
            st.markdown(f"**Yesterday ({yesterday})**")
            components.html(
                get_activity_block_html(yesterday),
                height=260,
                scrolling=True
            )

        # ---- TODAY ----
        with col_t:
            st.markdown(f"**Today ({today})**")
            components.html(
                get_activity_block_html(today),
                height=260,
                scrolling=True
            )

        # ---- TOMORROW ----
        with col_tm:
            st.markdown(f"**Tomorrow ({tomorrow})**")
            components.html(
                get_activity_block_html(tomorrow),
                height=260,
                scrolling=True
            )

        st.divider()

        # ---- LOG ENTRY SECTION ----
        project = st.selectbox("Project", proj_names, key="log_project")
        engineer = st.selectbox("Engineer", engineers, key="log_engineer")
        st.info(f"IFS Code: {proj_map.get(project, 'Not Set')}")

        hours = st.number_input("Hours", 0.0, 24.0, step=0.25, key="log_hours")
        note = st.text_area("Activity Note", key="log_note")

        if st.button("✅ Submit", key="log_submit_btn"):
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Log"]
            now = datetime.now()

            ws.append([
                base_date.strftime("%Y-%m-%d"),
                now.strftime("%H:%M:%S"),
                project,
                engineer,
                hours,
                note
            ])
            ws[f"F{ws.max_row}"].alignment = Alignment(wrap_text=True)

            wb.save(EXCEL_FILE)
            wb.close()

            st.success("✅ Activity saved")


            st.rerun()

    # ---------- TODAY ----------
    with tab_editactivity:
        st.subheader("📅 View / Edit / Add Activities")

        # -------------------------------------------------
        # Session State Initialization
        # -------------------------------------------------
        defaults = {
            "searched": False,
            "search_date": date.today(),
            "preview_df": None,
            "edit_mode": False,
            "add_mode": False,
        }
        for k, v in defaults.items():
            st.session_state.setdefault(k, v)





        # -------------------------------------------------
        # Date Selection
        # -------------------------------------------------
        selected_date = st.date_input(
            "Select Date",
            st.session_state.search_date,
            key="edit_activity_select_date"
        )

        if st.button("🔍 Search", key="edit_activity_search_btn"):
            st.session_state.searched = True
            st.session_state.search_date = selected_date
            st.session_state.edit_mode = False
            st.session_state.add_mode = False

        # -------------------------------------------------
        # Load data
        # -------------------------------------------------
        if st.session_state.searched:
            wb = load_workbook(EXCEL_FILE, read_only=True)
            ws = wb["Log"]

            rows = []
            for row_no, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if parse_excel_date(r[0]) == st.session_state.search_date:
                    rows.append({
                        "row_no": row_no,
                        "time": r[1],
                        "project": r[2],
                        "engineer": r[3],
                        "hours": float(r[4]) if r[4] else 0.0,
                        "note": r[5] or ""
                    })
            wb.close()

            st.session_state.preview_df = pd.DataFrame(rows) if rows else None

        # -------------------------------------------------
        # PREVIEW
        # -------------------------------------------------
        if st.session_state.preview_df is not None:
            st.subheader("📋 Activity Preview")

            render_preview_with_multiline_notes(
                st.session_state.preview_df.drop(columns=["row_no"]),
                height_px=300
            )

            st.success(
                f"✅ Total Hours = {st.session_state.preview_df['hours'].sum():.2f}"
            )

            col1, col2 = st.columns(2)
            if col1.button("✏ Edit Activities", key="edit_activities_btn"):
                st.session_state.edit_mode = True
                st.session_state.add_mode = False
            if col2.button("➕ Add Activity", key="add_activity_btn_preview"):
                st.session_state.add_mode = True
                st.session_state.edit_mode = False

        elif st.session_state.searched:
            st.warning("No activities found for selected date.")
            if st.button("➕ Add Activity", key="add_activity_btn_no_data"):
                st.session_state.add_mode = True

        # -------------------------------------------------
        # EDIT ACTIVITIES (FORM-BASED, MULTILINE SAFE)
        # -------------------------------------------------
        if st.session_state.edit_mode and st.session_state.preview_df is not None:
            st.subheader("✏ Edit Activities")

            for i, r in st.session_state.preview_df.iterrows():
                with st.expander(
                        f"{r['time']} | {r['project']} | {r['engineer']} | {r['hours']} h"
                ):
                    new_hours = st.number_input(
                        "Hours",
                        min_value=0.0,
                        max_value=24.0,
                        step=0.25,
                        value=float(r["hours"]),
                        key=f"hours_{i}"
                    )

                    new_note = st.text_area(
                        "Activity Note",
                        value=r["note"],
                        height=120,
                        key=f"note_{i}"
                    )

                    # ---- SAVE CHANGES ----
                    if st.button("💾 Save Changes", key=f"save_{i}"):
                        wb = load_workbook(EXCEL_FILE)
                        ws = wb["Log"]

                        excel_row = r["row_no"]
                        ws.cell(excel_row, 5).value = new_hours
                        ws.cell(excel_row, 6).value = new_note
                        ws.cell(excel_row, 6).alignment = Alignment(wrap_text=True)

                        wb.save(EXCEL_FILE)
                        wb.close()

                        st.success("✅ Activity updated")
                        st.session_state.edit_mode = False
                        st.rerun()

                    st.divider()

                    # ---- DELETE ACTIVITY ----
                    confirm_delete = st.checkbox(
                        "⚠️ Yes, I want to delete this activity",
                        key=f"confirm_delete_{i}"
                    )

                    if confirm_delete:
                        if st.button("❌ Delete Activity", key=f"delete_{i}"):
                            wb = load_workbook(EXCEL_FILE)
                            ws = wb["Log"]

                            ws.delete_rows(r["row_no"], 1)

                            wb.save(EXCEL_FILE)
                            wb.close()

                            st.error("❌ Activity deleted")
                            st.session_state.edit_mode = False
                            st.rerun()

        # -------------------------------------------------
        # ADD NEW ACTIVITY
        # -------------------------------------------------
        if st.session_state.add_mode:
            st.subheader("➕ Add New Activity")

            with st.form("add_activity_form", clear_on_submit=True):
                col1, col2 = st.columns(2)
                project = col1.selectbox("Project", proj_names)
                engineer = col2.selectbox("Engineer", engineers)
                hours = st.number_input("Hours", 0.0, 24.0, step=0.25)
                note = st.text_area("Activity Note")
                submitted = st.form_submit_button("✅ Save Activity")

            if submitted:
                wb = load_workbook(EXCEL_FILE)
                ws = wb["Log"]
                now = datetime.now()

                ws.append([
                    st.session_state.search_date.strftime("%Y-%m-%d"),
                    now.strftime("%H:%M:%S"),
                    project,
                    engineer,
                    hours,
                    note
                ])
                ws[f"F{ws.max_row}"].alignment = Alignment(wrap_text=True)

                wb.save(EXCEL_FILE)
                wb.close()

                st.success("✅ Activity added")
                st.session_state.add_mode = False
                st.rerun()

    # ---------- WEEK ----------
    with tab_summary:
        # Generate Summary
        st.subheader("Generate Summary")

        # ---- Date Range ----
        col1, col2 = st.columns(2)

        # Initialize end date only once
        if "summary_end_date" not in st.session_state:
            st.session_state.summary_end_date = date.today()

        start = col1.date_input("Start Date")

        end = col2.date_input(
            "End Date",
            value=st.session_state.summary_end_date,
            key="summary_end_date"
        )

        # ---- Summary Type ----
        summary_type = st.radio(
            "Summary Type",
            ["Date-wise", "Engineer-wise", "Project-wise"],
            horizontal=True
        )

        st.divider()

        # ---- Filters ----
        pf, ef = st.columns(2)
        selected_projects = pf.multiselect(
            "Select Project(s)",
            proj_names,
            default=proj_names
        )
        selected_engineers = ef.multiselect(
            "Select Engineer(s)",
            engineers,
            default=engineers
        )

        st.divider()

        # ---- Generate Summary ----
        if st.button("📊 Generate Summary"):
            wb = load_workbook(EXCEL_FILE, read_only=True)
            ws = wb["Log"]
            records = []

            for r in ws.iter_rows(min_row=2, values_only=True):
                d = parse_excel_date(r[0])
                if not d or not (start <= d <= end):
                    continue

                if r[2] not in selected_projects:
                    continue
                if r[3] not in selected_engineers:
                    continue

                records.append({
                    "date": d,
                    "time": r[1][:5] if r[1] else "--:--",
                    "project": r[2],
                    "engineer": r[3],
                    "hours": float(r[4]) if r[4] else 0.0,
                    "note": r[5] or ""
                })

            wb.close()
            st.session_state.weekly = records

        records = st.session_state.get("weekly", [])

        if records:
            preview_df = pd.DataFrame(records)
        if records:
            preview_df["date"] = preview_df["date"].astype(str)
            preview_df = preview_df[
                ["date", "project", "engineer", "hours", "note"]
            ]
        if not records:
            st.info("No data found for selected filters.")
        else:
            st.subheader("📋 Summary Overview")

            preview_df = pd.DataFrame(records)
            preview_df["date"] = preview_df["date"].astype(str)
            preview_df = preview_df.sort_values(["date", "time"])

            preview_df["date"] = preview_df["date"].where(
                preview_df["date"].ne(preview_df["date"].shift()),
                ""
            )

            # ---- Hide repeating Project (ONLY within same Date) ----
            preview_df["project"] = preview_df["project"].where(
                ~(
                        (preview_df["project"] == preview_df["project"].shift()) &
                        (preview_df["date"] == "")
                ),
            )

            # ✅ REMOVE TIME COLUMN
            preview_df = preview_df[
                ["date", "project", "engineer", "hours", "note"]
            ]

            # ✅ SENTENCE CASE HEADERS
            preview_df = preview_df.rename(columns={
                "date": "Date",
                "project": "Project",
                "engineer": "Engineer",
                "hours": "Hours",
                "note": "Note"
            })

            col_left, col_right = st.columns([1, 2])

            # -------- LEFT : TIME ENTRY CODE VIEW --------
            with col_left:
                st.markdown("### ⏱ Time Entry Code View")
                components.html(
                    render_time_entry_code_view_html(records, proj_map),
                    height=360,
                    scrolling=True
                )

            # -------- RIGHT : SUMMARY PREVIEW --------
            with col_right:
                st.markdown("### 📄 Summary Preview")
                render_preview_with_multiline_notes(
                    preview_df,
                    height_px=360
                )

            st.caption(
                "Left: compact time‑booking view | Right: detailed activity preview."
            )

            st.divider()

            # =====================================================
            # DATE‑WISE VIEW
            # =====================================================
            if summary_type == "Date-wise":
                grouped = {}
                for r in records:
                    grouped.setdefault(r["date"], []).append(r)

                for d, items in sorted(grouped.items()):
                    day_total = sum(x["hours"] for x in items)
                    with st.expander(f"{d} — Total {day_total:.2f} h"):
                        items.sort(key=lambda x: x["time"])
                        for x in items:
                            st.write(
                                f"⏱ {x['time']}  |  "
                                f"📌 {x['project']}  |  "
                                f"👷 {x['engineer']}  |  "
                                f"⏳ {x['hours']:.2f} h"
                            )
                            st.caption(
                                f"📝 {x['note']}" if x["note"] else "📝 No notes provided"
                            )
                            st.divider()

            # =====================================================
            # ENGINEER‑WISE VIEW
            # =====================================================
            elif summary_type == "Engineer-wise":
                grouped = {}
                for r in records:
                    grouped.setdefault(r["engineer"], []).append(r)

                for engineer, items in grouped.items():
                    eng_total = sum(x["hours"] for x in items)
                    with st.expander(f"{engineer} — Total {eng_total:.2f} h"):
                        for x in items:
                            st.write(
                                f"🗓 {x['date']}  |  "
                                f"⏱ {x['time']}  |  "
                                f"📌 {x['project']}  |  "
                                f"⏳ {x['hours']:.2f} h"
                            )
                            st.caption(
                                f"📝 {x['note']}" if x["note"] else "📝 No notes provided"
                            )
                            st.divider()

            # =====================================================
            # PROJECT‑WISE VIEW
            # =====================================================
            elif summary_type == "Project-wise":
                grouped = {}
                for r in records:
                    grouped.setdefault(r["project"], []).append(r)

                for project, items in grouped.items():
                    proj_total = sum(x["hours"] for x in items)
                    with st.expander(f"{project} — Total {proj_total:.2f} h"):
                        for x in items:
                            st.write(
                                f"🗓 {x['date']}  |  "
                                f"⏱ {x['time']}  |  "
                                f"👷 {x['engineer']}  |  "
                                f"⏳ {x['hours']:.2f} h"
                            )
                            st.caption(
                                f"📝 {x['note']}" if x["note"] else "📝 No notes provided"
                            )
                            st.divider()

            # ---- ACTIONS (ONLY ONCE) ----
            st.divider()

            if st.button("⬇ Download Excel", key="download_week_excel"):
                wb = create_weekly_summary_excel(records)
                buffer = BytesIO()
                wb.save(buffer)
                wb.close()
                buffer.seek(0)

                st.download_button(
                    label="📥 Download Weekly Summary",
                    data=buffer,
                    file_name="Weekly_Summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_week_excel_btn"
                )

            if st.button("📄 Save Summary to Activity Excel", key="save_week_summary"):
                append_weekly_summary(records)
                st.success("Weekly summary sheet created")

    # ---------- SETTINGS ----------
    # ---------- SETTINGS ----------
    if st.session_state.role == "Administrator":
        with tab_settings:

            # ---------------- PROJECTS & ENGINEERS ----------------
            col1, col2 = st.columns(2)
            with col1:
                manage_projects_wizard()
            with col2:
                manage_engineers_wizard()

            st.divider()

            # ---------------- REMINDER SETTINGS ----------------
            popup_min = st.number_input(
                "Reminder interval (minutes)",
                min_value=1,
                value=settings["popup_min"]
            )

            if st.button("Save Settings", key="save_popup_settings"):
                settings["popup_min"] = popup_min
                json.dump(settings, open(SETTINGS_FILE, "w"), indent=2)
                st.success("✅ Settings saved")

            st.divider()

            # ---------------- USER MANAGEMENT ----------------
            st.subheader("🔐 Manage Users (Login Access)")

            users_df = pd.DataFrame.from_dict(users_db, orient="index")
            users_df.index.name = "Username"
            users_df.reset_index(inplace=True)

            edited_users = st.data_editor(
                users_df,
                use_container_width=True,
                num_rows="dynamic",
                column_config={
                    "role": st.column_config.SelectboxColumn(
                        "Role",
                        options=["Administrator", "User"]
                    ),
                    "locked": st.column_config.CheckboxColumn("Locked")
                }
            )

            if st.button("💾 Save Users", key="save_users"):
                updated_users = {}

                for _, row in edited_users.iterrows():
                    uname = row["Username"].strip()
                    if not uname:
                        continue

                    # ✅ OPTIONAL: prevent self-lock
                    if uname == st.session_state.username and row["locked"]:
                        st.warning("⚠️ You cannot lock your own account")
                        continue

                    updated_users[uname] = {
                        "password": row["password"],
                        "role": row["role"],
                        "locked": bool(row["locked"])
                    }

                json.dump(updated_users, open(USERS_FILE, "w"), indent=2)
                st.success("✅ Users updated successfully")
                st.rerun()

            st.divider()

            # ---------------- PASSWORD RESET ----------------
            st.subheader("🔑 Reset User Password")

            reset_user = st.selectbox(
                "Select User",
                list(users_db.keys()),
                key="reset_user_select"
            )
            new_password = st.text_input(
                "New Password",
                type="password",
                key="reset_pwd_input"
            )

            if st.button("Reset Password", key="reset_pwd_btn"):
                if new_password.strip():
                    users_db[reset_user]["password"] = new_password
                    json.dump(users_db, open(USERS_FILE, "w"), indent=2)
                    st.success(f"✅ Password reset for {reset_user}")
                else:
                    st.warning("Password cannot be empty")


if page == "Page-2":
    tab_1, tab_2, tab_3, tab_4 = st.tabs(
        ["📝 Tab-01", "📝 Tab-01", "📝 Tab-01", "📝 Tab-01"]
    )